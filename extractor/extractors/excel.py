"""
extractors/excel.py
Extracts data from Excel (.xlsx) workbooks.
Each sheet becomes one "page" in the output.

Install:
    pip install openpyxl
"""


def extract_excel(file_path: str) -> list:
    """
    Reads every sheet in the workbook.
    Converts each row to a pipe-delimited string so the LLM can parse it.
    Skips completely empty rows.

    Returns:
        List of dicts — one per sheet — each with keys:
        page, text, source, type, method, sheet_name, has_tables
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl not installed.\n"
            "Run: pip install openpyxl"
        )

    # data_only=True reads computed cell values, not formulas
    wb      = openpyxl.load_workbook(file_path, data_only=True)
    source  = file_path.split("/")[-1]
    results = []

    for sheet_num, sheet_name in enumerate(wb.sheetnames, start=1):
        ws   = wb[sheet_name]
        rows = []

        for row in ws.iter_rows(values_only=True):
            # Convert each cell to string, replace None with empty string
            clean = [str(c).strip() if c is not None else "" for c in row]
            # Skip rows where every cell is empty
            if any(c for c in clean):
                rows.append(" | ".join(clean))

        sheet_text = f"[SHEET: {sheet_name}]\n" + "\n".join(rows)

        results.append({
            "page":       sheet_num,
            "text":       sheet_text.strip(),
            "source":     source,
            "type":       "xlsx",
            "method":     "openpyxl",
            "sheet_name": sheet_name,
            "has_tables": True,
        })

    return results